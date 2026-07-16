"""
overtime_summary_engine.py
============================
Calculation / Excel-building logic for the "OT & Attendance Summary" sheet.

This is a COMPLETELY SEPARATE engine from timesheet_engine.py — it does not
import from it and does not share any state. It has its own file
classifier, its own dataframe loader, and its own workbook builder, so the
two features can never conflict or interfere with each other.

Required input files (any order, any file names — auto-detected by columns):
  1. Attendance / Punches file   — needs a column named "I/O"
  2. Vacation Transaction file   — needs columns "Vacation" + "From"
  3. Employee master file        — needs columns "Employees Name" + "Title"
     (reads the FIRST sheet — e.g. the "Data" sheet in Data_2026.xlsx)
  4. Official Holidays file      — needs columns "التاريخ" (Date) + "المناسبة" (Occasion)

Matching between files is done by employee Code (not Name).
"""

import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config defaults ────────────────────────────────────────────────────────
DEFAULT_WORKDAY_HRS = 8.0
DEFAULT_WEEKEND_DAYS = {4, 5}  # Friday=4, Saturday=5


class OTSummaryError(Exception):
    """Raised for any user-facing problem (missing columns, bad files, etc.)."""
    pass


# ── Step 1: classify the 4 uploaded files ───────────────────────────────────
def classify_files(file_dict):
    """
    file_dict: {filename: file-like object}, exactly 4 files.
    Returns (att_file, vac_file, emp_file, hol_file), each seeked to 0.
    """
    att_file = vac_file = emp_file = hol_file = None
    problems = []

    for fname, fobj in file_dict.items():
        try:
            fobj.seek(0)
            df_peek = pd.read_excel(fobj, nrows=3)
            df_peek.columns = [str(c).strip() for c in df_peek.columns]
            cols_lower = [c.lower() for c in df_peek.columns]
            cols_raw = list(df_peek.columns)

            if "i/o" in cols_lower:
                att_file = fobj
            elif "vacation" in cols_lower and "from" in cols_lower:
                vac_file = fobj
            elif "التاريخ" in cols_raw or "المناسبة" in cols_raw:
                hol_file = fobj
            elif "employees name" in cols_lower and "title" in cols_lower:
                emp_file = fobj
        except Exception as e:
            problems.append(f"Could not read '{fname}': {e}")
        finally:
            fobj.seek(0)

    missing = []
    if not att_file:
        missing.append("Attendance/Punches file — needs a column named 'I/O'")
    if not vac_file:
        missing.append("Vacation Transaction file — needs columns 'Vacation' + 'From'")
    if not emp_file:
        missing.append("Employee master file — needs columns 'Employees Name' + 'Title'")
    if not hol_file:
        missing.append("Official Holidays file — needs a date column 'التاريخ' and 'المناسبة'")

    if missing:
        msg = "Could not identify all 4 required files:\n- " + "\n- ".join(missing)
        if problems:
            msg += "\n\nAlso had trouble reading some files:\n- " + "\n- ".join(problems)
        raise OTSummaryError(msg)

    return att_file, vac_file, emp_file, hol_file


# ── Step 2: normalize a Code value into a consistent string key ────────────
def norm_code(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s if s else None


def _find_col(df, *candidates):
    """Case-insensitive column finder. Returns the actual column name or None."""
    lower_map = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


# ── Step 3: load & clean the 4 dataframes ───────────────────────────────────
def load_dataframes(att_file, vac_file, emp_file, hol_file):
    df_att = pd.read_excel(att_file)
    df_vac = pd.read_excel(vac_file)
    df_emp = pd.read_excel(emp_file)  # first sheet only
    df_hol = pd.read_excel(hol_file)

    df_att.columns = [str(c).strip() for c in df_att.columns]
    df_vac.columns = [str(c).strip() for c in df_vac.columns]
    df_emp.columns = [str(c).strip() for c in df_emp.columns]
    df_hol.columns = [str(c).strip() for c in df_hol.columns]

    # ---- Attendance ----
    code_col = _find_col(df_att, "Code")
    date_col = _find_col(df_att, "Date")
    time_col = _find_col(df_att, "Time")
    io_col = _find_col(df_att, "I/O")
    if not all([code_col, date_col, time_col, io_col]):
        raise OTSummaryError("Attendance file is missing one of: Code, Date, Time, I/O columns.")

    df_att["Date"] = pd.to_datetime(df_att[date_col], errors="coerce").dt.date
    df_att["Time"] = pd.to_datetime(df_att[time_col], errors="coerce").dt.time
    df_att["CodeKey"] = df_att[code_col].map(norm_code)
    df_att["IO"] = df_att[io_col]
    df_att = df_att.dropna(subset=["Date", "CodeKey"])

    if df_att.empty:
        raise OTSummaryError(
            "The Attendance file has no valid rows after cleaning. "
            "Check that 'Date' and 'Code' columns are filled in correctly."
        )

    # ---- Vacation (ANY status / ANY type counts as an excuse) ----
    v_code = _find_col(df_vac, "Code")
    v_from = _find_col(df_vac, "From")
    v_to = _find_col(df_vac, "To")
    if not all([v_code, v_from, v_to]):
        raise OTSummaryError("Vacation file is missing one of: Code, From, To columns.")

    df_vac["From"] = pd.to_datetime(df_vac[v_from], errors="coerce").dt.date
    df_vac["To"] = pd.to_datetime(df_vac[v_to], errors="coerce").dt.date
    df_vac["CodeKey"] = df_vac[v_code].map(norm_code)
    df_vac = df_vac.dropna(subset=["From", "To", "CodeKey"])

    # ---- Employee master ----
    e_code = _find_col(df_emp, "Code", "code")
    e_name = _find_col(df_emp, "Employees Name")
    e_title = _find_col(df_emp, "Title")
    e_dept = _find_col(df_emp, "Department")
    if not all([e_code, e_name]):
        raise OTSummaryError("Employee master file is missing 'Code' or 'Employees Name' columns.")

    df_emp["CodeKey"] = df_emp[e_code].map(norm_code)
    df_emp["Employees Name"] = df_emp[e_name].astype(str).str.strip()
    df_emp["Title_"] = df_emp[e_title].astype(str).str.strip() if e_title else ""
    df_emp["Dept_"] = df_emp[e_dept].astype(str).str.strip() if e_dept else ""

    df_emp_master = df_emp[
        df_emp["CodeKey"].notna()
        & df_emp["CodeKey"].str.match(r"^\d+$", na=False)  # numeric codes only (skips '*' rows)
        & df_emp["Employees Name"].notna()
        & (df_emp["Employees Name"] != "")
        & (df_emp["Employees Name"].str.lower() != "nan")
    ].drop_duplicates(subset=["CodeKey"], keep="first").reset_index(drop=True)

    if df_emp_master.empty:
        raise OTSummaryError(
            "The Employee master file has no valid numeric-coded rows after cleaning."
        )

    emp_by_code = df_emp_master.set_index("CodeKey").to_dict("index")

    # ---- Official Holidays ----
    h_date = _find_col(df_hol, "التاريخ") or df_hol.columns[0]
    h_occ = _find_col(df_hol, "المناسبة")
    df_hol["HolDate"] = pd.to_datetime(df_hol[h_date], errors="coerce", dayfirst=True).dt.date
    df_hol = df_hol.dropna(subset=["HolDate"])
    holiday_dates = set(df_hol["HolDate"].tolist())
    holiday_names = {
        row["HolDate"]: str(row[h_occ]).strip() if h_occ else ""
        for _, row in df_hol.iterrows()
    }

    return df_att, df_vac, emp_by_code, holiday_dates, holiday_names


# ── Step 4: per-employee helpers ────────────────────────────────────────────
def get_daily_records(df_att, code_key):
    """Returns {date: (earliest_in, latest_out)} for one employee."""
    person_df = df_att[df_att["CodeKey"] == code_key]
    records = {}
    for date, g in person_df.groupby("Date"):
        ins = g[g["IO"] == "Punch In"]["Time"].dropna().tolist()
        outs = g[g["IO"] == "Punch Out"]["Time"].dropna().tolist()
        t_in = min(ins) if ins else None
        t_out = max(outs) if outs else None
        records[date] = (t_in, t_out)
    return records


def _combine(date, t):
    return datetime.datetime(date.year, date.month, date.day, t.hour, t.minute, t.second)


def calc_hours(date, t_in, t_out):
    """Worked hours for the day; handles overnight punch-out."""
    if not t_in or not t_out:
        return None
    a = _combine(date, t_in)
    b = _combine(date, t_out)
    if b < a:
        b += datetime.timedelta(days=1)
    return round((b - a).total_seconds() / 3600, 2)


def split_ot_hours(ot_start_dt, ot_end_dt):
    """
    Splits the overtime interval [ot_start_dt, ot_end_dt) into:
      - morning OT hours: portion falling within 07:00–19:00
      - night OT hours:   portion falling within 19:00–07:00 (wraps midnight)
    Returns (morning_hours, night_hours).
    """
    morning_secs = 0.0
    night_secs = 0.0
    cur = ot_start_dt
    while cur < ot_end_dt:
        day_start = datetime.datetime(cur.year, cur.month, cur.day)
        morning_start = day_start + datetime.timedelta(hours=7)
        morning_end = day_start + datetime.timedelta(hours=19)
        next_day_start = day_start + datetime.timedelta(days=1)
        seg_end = min(ot_end_dt, next_day_start)

        overlap_start = max(cur, morning_start)
        overlap_end = min(seg_end, morning_end)
        m_secs = (overlap_end - overlap_start).total_seconds() if overlap_start < overlap_end else 0.0

        seg_total_secs = (seg_end - cur).total_seconds()
        morning_secs += m_secs
        night_secs += (seg_total_secs - m_secs)

        cur = seg_end

    return round(morning_secs / 3600, 2), round(night_secs / 3600, 2)


def is_excused(df_vac, code_key, target_date):
    """True if ANY vacation record (any type, any status) covers this date."""
    rows = df_vac[df_vac["CodeKey"] == code_key]
    for _, row in rows.iterrows():
        if row["From"] <= target_date <= row["To"]:
            return True
    return False


def emp_info(emp_by_code, code_key):
    data = emp_by_code.get(code_key, {})
    code = str(data.get("code", code_key)).strip() if "code" in data else code_key
    # fall back to whatever the original code column value was
    name = str(data.get("Employees Name", "")).strip()
    title = str(data.get("Title_", "")).strip()
    dept = str(data.get("Dept_", "")).strip()
    title = title if title and title.lower() != "nan" else "—"
    dept = dept if dept and dept.lower() != "nan" else "—"
    return code_key, name, title, dept


# ── Step 5: styles ───────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="D0D5DD")
_BDR = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FILL = PatternFill("solid", fgColor="0F766E")   # teal — visually distinct from Timesheet
_HEADER_FILL = PatternFill("solid", fgColor="CCFBF1")
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
_TITLE_FONT_COLOR = "FFFFFF"


def _sc(ws, row, col, value=None, bold=False, fill=None, align="center", fmt=None, size=10, color="1F2937"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=size, color=color)
    c.border = _BDR
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


# ── Step 6: build the workbook ────────────────────────────────────────────────
def generate_ot_summary(
    file_dict,
    workday_hrs=DEFAULT_WORKDAY_HRS,
    weekend_days=None,
    progress_callback=None,
):
    """
    file_dict: {filename: file-like object}, exactly 4 files.
    Returns: (workbook, output_filename, stats)
    """
    if weekend_days is None:
        weekend_days = DEFAULT_WEEKEND_DAYS

    if len(file_dict) != 4:
        raise OTSummaryError(f"Expected exactly 4 files, got {len(file_dict)}.")

    att_file, vac_file, emp_file, hol_file = classify_files(file_dict)
    df_att, df_vac, emp_by_code, holiday_dates, holiday_names = load_dataframes(
        att_file, vac_file, emp_file, hol_file
    )

    sys_min = df_att["Date"].min()
    sys_max = df_att["Date"].max()
    if pd.isna(sys_min) or pd.isna(sys_max):
        raise OTSummaryError("Could not find any valid dates in the Attendance file's 'Date' column.")

    all_dates = [sys_min + datetime.timedelta(days=i) for i in range((sys_max - sys_min).days + 1)]
    period_holidays = {d for d in holiday_dates if sys_min <= d <= sys_max}

    roster = sorted(emp_by_code.items(), key=lambda kv: kv[1]["Employees Name"])

    month_str = (
        f"{sys_min.strftime('%B')} – {sys_max.strftime('%B %Y')}"
        if sys_min.month != sys_max.month
        else sys_min.strftime("%B %Y")
    )
    output_file = f"OT_Attendance_Summary_{sys_min.strftime('%d%b')}_{sys_max.strftime('%d%b_%Y')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "OT & Attendance Summary"

    headers = [
        "Code", "Name", "Title / Position", "Department",
        "Morning OT Hours", "Night OT Hours",
        "Cancel Day Offs / Days", "Official Holiday / Days",
        "Unpaid Leave", "Total Working Days",
    ]
    col_widths = [10, 26, 22, 20, 16, 16, 18, 18, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    _sc(ws, 1, 1, f"OT & Attendance Summary  |  {month_str}", bold=True, fill=_TITLE_FILL,
        size=13, color=_TITLE_FONT_COLOR)
    for col in range(2, len(headers) + 1):
        ws.cell(row=1, column=col).fill = _TITLE_FILL
        ws.cell(row=1, column=col).border = _BDR
    ws.row_dimensions[1].height = 24

    # Header row
    for col, hdr in enumerate(headers, 1):
        _sc(ws, 2, col, hdr, bold=True, fill=_HEADER_FILL)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    r = 3
    total_roster = len(roster)
    for emp_idx, (code_key, _info) in enumerate(roster, 1):
        _, name, title, dept = emp_info(emp_by_code, code_key)
        records = get_daily_records(df_att, code_key)

        morning_ot = 0.0
        night_ot = 0.0
        cancel_day_off_days = 0.0
        official_holiday_days = 0.0
        unpaid_leave_days = 0
        total_working_days = 0

        for date in all_dates:
            t_in, t_out = records.get(date, (None, None))
            hrs = calc_hours(date, t_in, t_out)
            is_off_day = (date.weekday() in weekend_days) or (date in period_holidays)

            if t_in is not None:
                total_working_days += 1

            if is_off_day:
                if hrs and hrs > 0:
                    if hrs < workday_hrs:
                        cancel_day_off_days += round(hrs / workday_hrs, 2)
                    else:
                        official_holiday_days += 1.0
                continue  # off-days never generate Morning/Night OT or Unpaid Leave

            # Regular workday
            if hrs and hrs > workday_hrs:
                ot_start = _combine(date, t_in) + datetime.timedelta(hours=workday_hrs)
                ot_end = _combine(date, t_out)
                if ot_end < ot_start:
                    ot_end += datetime.timedelta(days=1)
                m_ot, n_ot = split_ot_hours(ot_start, ot_end)
                morning_ot += m_ot
                night_ot += n_ot

            if t_in is None and t_out is None:
                if not is_excused(df_vac, code_key, date):
                    unpaid_leave_days += 1

        fill = _ALT_FILL if emp_idx % 2 == 0 else None
        _sc(ws, r, 1, code_key, fill=fill)
        _sc(ws, r, 2, name, fill=fill, align="left")
        _sc(ws, r, 3, title, fill=fill, align="left")
        _sc(ws, r, 4, dept, fill=fill, align="left")
        _sc(ws, r, 5, round(morning_ot, 2) if morning_ot else 0, fill=fill, fmt="0.00")
        _sc(ws, r, 6, round(night_ot, 2) if night_ot else 0, fill=fill, fmt="0.00")
        _sc(ws, r, 7, round(cancel_day_off_days, 2) if cancel_day_off_days else 0, fill=fill, fmt="0.00")
        _sc(ws, r, 8, round(official_holiday_days, 2) if official_holiday_days else 0, fill=fill, fmt="0.00")
        _sc(ws, r, 9, unpaid_leave_days, fill=fill)
        _sc(ws, r, 10, total_working_days, fill=fill)
        r += 1

        if progress_callback:
            progress_callback(emp_idx, total_roster)

    stats = {
        "num_employees": len(roster),
        "num_days": len(all_dates),
        "month_str": month_str,
        "num_holidays_in_period": len(period_holidays),
    }
    return wb, output_file, stats
