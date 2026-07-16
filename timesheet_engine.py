"""
timesheet_engine.py
====================
SECTION 1 — calculation engine for the Attendance Timesheet Generator.
All the calculation / Excel-building logic. This file has NO UI code and NO
Colab code — it just takes in-memory file data and returns a finished
workbook.

Matching is done by **Code** (employee ID), not by Name, because names are
spelled slightly differently across the System / Vacation / Employees Data
files (extra spaces, typos, transliteration differences).

NOTE: This file is intentionally independent from ot_payroll_engine.py
(Section 2). Do not import from it and do not let Section 2 changes touch
this file — that's what keeps the two sections from ever breaking each other.
"""

import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config defaults (can be overridden by the UI) ────────────────────────────
DEFAULT_WORKDAY_HRS = 8.0
DEFAULT_WORK_START = "10:00"
DEFAULT_WEEKEND_DAYS = {4, 5}  # Friday=4, Saturday=5 (Python: Mon=0 ... Sun=6)


class TimesheetError(Exception):
    """Raised for any user-facing problem (missing columns, bad files, etc.)."""
    pass


# ── Step 1: figure out which uploaded file is which ──────────────────────────
def classify_files(file_dict):
    """
    file_dict: {filename: file-like object} (e.g. from Streamlit's uploader,
    or open("path", "rb")). Each file-like object must support pandas'
    read_excel (i.e. be seekable).

    Returns (system_file, emp_file, vac_file) as the same file-like objects
    that were passed in, seeked back to position 0.

    Raises TimesheetError if any of the 3 roles can't be found.
    """
    system_file = emp_file = vac_file = None
    problems = []

    for fname, fobj in file_dict.items():
        try:
            fobj.seek(0)
            df_peek = pd.read_excel(fobj, nrows=3)
            df_peek.columns = df_peek.columns.str.strip()
            cols = [c.lower() for c in df_peek.columns]
            if "i/o" in cols:
                system_file = fobj
            elif "title" in cols and "employees name" in cols:
                emp_file = fobj
            elif "vacation" in cols and "from" in cols:
                vac_file = fobj
        except Exception as e:
            problems.append(f"Could not read '{fname}': {e}")
        finally:
            fobj.seek(0)

    missing = []
    if not system_file:
        missing.append("Attendance/System file — needs a column named 'I/O'")
    if not emp_file:
        missing.append("Employees Data file — needs columns 'Employees Name' + 'Title'")
    if not vac_file:
        missing.append("Vacation Transaction file — needs columns 'Vacation' + 'From'")

    if missing:
        msg = "Could not identify all 3 required files:\n- " + "\n- ".join(missing)
        if problems:
            msg += "\n\nAlso had trouble reading some files:\n- " + "\n- ".join(problems)
        raise TimesheetError(msg)

    return system_file, emp_file, vac_file


# ── Step 2: normalize a Code value into a consistent string key ──────────────
def norm_code(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s if s else None


# ── Step 3: load & clean the 3 dataframes ─────────────────────────────────────
def load_dataframes(system_file, emp_file, vac_file):
    df_sys = pd.read_excel(system_file)
    df_emp = pd.read_excel(emp_file)
    df_vac = pd.read_excel(vac_file)

    df_sys.columns = df_sys.columns.str.strip()
    df_emp.columns = df_emp.columns.str.strip()
    df_vac.columns = df_vac.columns.str.strip()

    # System file
    df_sys["Date"] = pd.to_datetime(df_sys["Date"], errors="coerce").dt.date
    df_sys["Time"] = pd.to_datetime(df_sys["Time"], errors="coerce").dt.time
    df_sys["CodeKey"] = df_sys["Code"].map(norm_code)
    df_sys = df_sys.dropna(subset=["Date", "CodeKey"])

    if df_sys.empty:
        raise TimesheetError(
            "The Attendance/System file has no valid rows after cleaning. "
            "Check that the 'Date' and 'Code' columns are filled in correctly."
        )

    # Vacation file
    df_vac["From"] = pd.to_datetime(df_vac["From"], errors="coerce").dt.date
    df_vac["To"] = pd.to_datetime(df_vac["To"], errors="coerce").dt.date
    df_vac["CodeKey"] = df_vac["Code"].map(norm_code)
    # Keep only real leave types (skip "-" which are partial-day permission records)
    df_vac_real = df_vac[df_vac["Vacation"].astype(str).str.strip() != "-"].copy()

    # Employee master: single source of truth for Code / Name / Title / Department
    df_emp["CodeKey"] = df_emp["Code"].map(norm_code)
    df_emp["Employees Name"] = df_emp["Employees Name"].astype(str).str.strip()
    df_emp_master = df_emp[
        df_emp["CodeKey"].notna()
        & df_emp["Employees Name"].notna()
        & (df_emp["Employees Name"] != "")
        & (df_emp["Employees Name"].str.lower() != "nan")
    ].drop_duplicates(subset=["CodeKey"], keep="first").reset_index(drop=True)

    if df_emp_master.empty:
        raise TimesheetError(
            "The Employees Data file has no valid rows after cleaning. "
            "Check the 'Code' and 'Employees Name' columns."
        )

    emp_by_code = df_emp_master.set_index("CodeKey").to_dict("index")

    return df_sys, df_vac_real, emp_by_code


# ── Step 4: per-employee helpers ──────────────────────────────────────────────
def get_daily_records(df_sys, code_key):
    """Returns {date: (earliest_in, latest_out)} for one employee, matched by Code."""
    person_df = df_sys[df_sys["CodeKey"] == code_key]
    records = {}
    for date, g in person_df.groupby("Date"):
        ins = g[g["I/O"] == "Punch In"]["Time"].dropna().tolist()
        outs = g[g["I/O"] == "Punch Out"]["Time"].dropna().tolist()
        records[date] = (min(ins) if ins else None, max(outs) if outs else None)
    return records


def calc_hours(t_in, t_out):
    if not t_in or not t_out:
        return None
    a = datetime.datetime(2000, 1, 1, t_in.hour, t_in.minute, t_in.second)
    b = datetime.datetime(2000, 1, 1, t_out.hour, t_out.minute, t_out.second)
    if b < a:
        b += datetime.timedelta(days=1)
    return round((b - a).seconds / 3600, 2)


def calc_delay(t_in, work_start):
    """Minutes late vs work_start; only if > 5 min."""
    if not t_in:
        return None
    expected = datetime.datetime.strptime(work_start, "%H:%M").replace(year=2000)
    actual = datetime.datetime(2000, 1, 1, t_in.hour, t_in.minute, t_in.second)
    if actual > expected:
        diff = int((actual - expected).seconds / 60)
        return diff if diff > 5 else None
    return None


def fmt_time(t):
    return t.strftime("%H:%M") if t else ""


def get_leave_type(df_vac_real, code_key, target_date):
    rows = df_vac_real[df_vac_real["CodeKey"] == code_key]
    for _, row in rows.iterrows():
        if pd.notna(row["From"]) and pd.notna(row["To"]):
            if row["From"] <= target_date <= row["To"]:
                return str(row["Vacation"]).strip()
    return ""


def emp_info(emp_by_code, code_key):
    data = emp_by_code.get(code_key, {})
    code = str(data.get("Code", code_key)).strip()
    name = str(data.get("Employees Name", "")).strip()
    title = str(data.get("Title", "")).strip()
    dept = str(data.get("Department", "")).strip()
    title = title if title and title.lower() != "nan" else "—"
    dept = dept if dept and dept.lower() != "nan" else "—"
    return code, name, title, dept


# ── Step 5: styles ─────────────────────────────────────────────────────────────
_THIN = Side(style="thin")
_BDR = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_WEEKEND_FILL = PatternFill("solid", fgColor="FCE4D6")
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5")
_TOTAL_FILL = PatternFill("solid", fgColor="D6DCE4")


def _sc(ws, row, col, value=None, bold=False, fill=None, align="center", fmt=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=size)
    c.border = _BDR
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


def _fill_row(ws, row, c1, c2, fill=None):
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = _BDR
        if fill:
            cell.fill = fill


# ── Step 6: build the workbook ────────────────────────────────────────────────
def generate_timesheet(
    file_dict,
    workday_hrs=DEFAULT_WORKDAY_HRS,
    work_start=DEFAULT_WORK_START,
    weekend_days=None,
    progress_callback=None,
):
    """
    file_dict: {filename: file-like object}, exactly 3 files.
    progress_callback: optional function(current, total) called after each
                        employee is processed, so a UI can show a progress bar.

    Returns: (workbook, output_filename, stats)
      stats = {"num_employees": int, "num_days": int, "month_str": str}
    """
    if weekend_days is None:
        weekend_days = DEFAULT_WEEKEND_DAYS

    if len(file_dict) != 3:
        raise TimesheetError(f"Expected exactly 3 files, got {len(file_dict)}.")

    system_file, emp_file, vac_file = classify_files(file_dict)
    df_sys, df_vac_real, emp_by_code = load_dataframes(system_file, emp_file, vac_file)

    sys_min = df_sys["Date"].min()
    sys_max = df_sys["Date"].max()
    if pd.isna(sys_min) or pd.isna(sys_max):
        raise TimesheetError("Could not find any valid dates in the System file's 'Date' column.")

    all_dates = [sys_min + datetime.timedelta(days=i) for i in range((sys_max - sys_min).days + 1)]

    roster = sorted(emp_by_code.items(), key=lambda kv: kv[1]["Employees Name"])

    month_str = (
        f"{sys_min.strftime('%B')} – {sys_max.strftime('%B %Y')}"
        if sys_min.month != sys_max.month
        else sys_min.strftime("%B %Y")
    )
    title_text = f"Staff Arabia Time Sheet  |  {month_str}"
    output_file = f"Timesheet_{sys_min.strftime('%d%b')}_{sys_max.strftime('%d%b_%Y')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    col_widths = {1: 14, 2: 14, 3: 10, 4: 10, 5: 13, 6: 12, 7: 14, 8: 26}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.sheet_format.defaultRowHeight = 15

    r = 1
    total_roster = len(roster)
    for emp_idx, (code_key, _info) in enumerate(roster, 1):
        e_code, name, e_title, e_dept = emp_info(emp_by_code, code_key)
        records = get_daily_records(df_sys, code_key)

        # Title row
        ws.merge_cells(f"A{r}:H{r}")
        _sc(ws, r, 1, title_text, bold=True, fill=_TITLE_FILL, size=12)
        _fill_row(ws, r, 2, 8, _TITLE_FILL)
        ws.row_dimensions[r].height = 20
        r += 1

        # Info rows
        for label, val in [
            ("Code :", e_code),
            ("Name :", name),
            ("Position :", e_title),
            ("Department :", e_dept),
        ]:
            ws.merge_cells(f"B{r}:H{r}")
            _sc(ws, r, 1, label, bold=True, align="left")
            _sc(ws, r, 2, val, bold=False, align="left")
            _fill_row(ws, r, 3, 8)
            r += 1

        # Column headers
        for col, hdr in enumerate(
            ["Date", "Day", "Time In", "Time Out", "Total Hours", "Overtime", "Delays (min)", "Leaves"], 1
        ):
            _sc(ws, r, col, hdr, bold=True, fill=_HEADER_FILL)
        r += 1

        # Data rows
        total_hrs = total_ot = total_delay = 0.0
        for i, date in enumerate(all_dates):
            is_weekend = date.weekday() in weekend_days
            fill = _WEEKEND_FILL if is_weekend else (_ALT_FILL if i % 2 == 0 else None)

            t_in, t_out = records.get(date, (None, None))
            hrs = calc_hours(t_in, t_out)
            ot = round(hrs - workday_hrs, 2) if hrs and hrs > workday_hrs else None
            delay = None if is_weekend else calc_delay(t_in, work_start)
            leave = get_leave_type(df_vac_real, code_key, date)

            if hrs:
                total_hrs += hrs
            if ot:
                total_ot += ot
            if delay:
                total_delay += delay

            _sc(ws, r, 1, date, fill=fill, fmt="dd-mmm-yy")
            _sc(ws, r, 2, date.strftime("%A"), fill=fill)
            _sc(ws, r, 3, fmt_time(t_in), fill=fill)
            _sc(ws, r, 4, fmt_time(t_out), fill=fill)
            _sc(ws, r, 5, hrs if hrs else "", fill=fill)
            _sc(ws, r, 6, ot if ot else "", fill=fill)
            _sc(ws, r, 7, delay if delay else "", fill=fill)
            _sc(ws, r, 8, leave, fill=fill)
            r += 1

        # Totals row
        _sc(ws, r, 1, "Total", bold=True, fill=_TOTAL_FILL)
        _fill_row(ws, r, 2, 8, _TOTAL_FILL)
        _sc(ws, r, 5, round(total_hrs, 2) if total_hrs else "", bold=True, fill=_TOTAL_FILL)
        _sc(ws, r, 6, round(total_ot, 2) if total_ot else "", bold=True, fill=_TOTAL_FILL)
        _sc(ws, r, 7, int(total_delay) if total_delay else "", bold=True, fill=_TOTAL_FILL)
        r += 1
        r += 3  # blank gap between employees

        if progress_callback:
            progress_callback(emp_idx, total_roster)

    stats = {
        "num_employees": len(roster),
        "num_days": len(all_dates),
        "month_str": month_str,
    }
    return wb, output_file, stats
