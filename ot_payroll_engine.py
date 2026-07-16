"""
ot_payroll_engine.py
======================
SECTION 2 — OT & Payroll Summary calculation engine.

This module is 100% independent from timesheet_engine.py (Section 1). It
does not import anything from it and nothing in Section 1 imports from here.
That separation is deliberate: changes to one section can never break the
other.

WHAT THIS COMPUTES (per employee, over the full date range found in the
attendance file):

- Morning OT Hours   : overtime minutes that fall inside 07:00–19:00
- Night OT Hours     : overtime minutes that fall inside 19:00–07:00
- Cancel Day Offs    : weekend/holiday days worked for < 8h, expressed as a
                       fraction of a day (hours / 8)
- Official Holiday   : weekend/holiday days worked for >= 8h, counted as 1
                       full day each
- Unpaid Leave       : regular workdays with NO punch at all AND NO vacation
                       record of any kind (any type, any status) covering
                       that date
- Total Working Days : distinct dates the employee has at least one punch-in

OT is only computed for *regular* (non-weekend, non-holiday) days. Weekend
and holiday days are routed entirely into the Cancel Day Off / Official
Holiday buckets instead, per the business rule.
"""

import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DEFAULT_WORKDAY_HRS = 8.0
DEFAULT_WEEKEND_DAYS = {4, 5}  # Friday=4, Saturday=5 (Python: Mon=0 ... Sun=6)

MORNING_START_MIN = 7 * 60   # 07:00
MORNING_END_MIN = 19 * 60    # 19:00


class OTPayrollError(Exception):
    """Raised for any user-facing problem (missing columns, bad files, etc.)."""
    pass


# ── Small shared helpers (deliberately duplicated from Section 1, NOT shared,
#    to keep this module fully self-contained) ────────────────────────────────
def norm_code(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s if s else None


def _find_col(columns, *candidates):
    """Case-insensitive lookup of the first matching column name, or None."""
    lower_map = {c.lower().strip(): c for c in columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


# ── Loading the required attendance file ──────────────────────────────────────
def load_attendance(file_obj):
    """
    Required columns (case-insensitive): Code, Name, Date, Time, I/O
    Returns a cleaned dataframe with CodeKey, Date (date), Time (time) columns.
    """
    file_obj.seek(0)
    df = pd.read_excel(file_obj)
    df.columns = df.columns.str.strip()

    code_col = _find_col(df.columns, "Code")
    name_col = _find_col(df.columns, "Name", "Employee Name", "Employees Name")
    date_col = _find_col(df.columns, "Date")
    time_col = _find_col(df.columns, "Time")
    io_col = _find_col(df.columns, "I/O", "IO")

    missing = [
        label
        for label, col in [("Code", code_col), ("Date", date_col), ("Time", time_col), ("I/O", io_col)]
        if col is None
    ]
    if missing:
        raise OTPayrollError(
            "The attendance file is missing required column(s): " + ", ".join(missing)
        )

    df = df.rename(columns={code_col: "Code", date_col: "Date", time_col: "Time", io_col: "I/O"})
    if name_col:
        df = df.rename(columns={name_col: "Name"})
    else:
        df["Name"] = ""

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    df["Time"] = pd.to_datetime(df["Time"], errors="coerce").dt.time
    df["CodeKey"] = df["Code"].map(norm_code)
    df = df.dropna(subset=["Date", "CodeKey"])

    if df.empty:
        raise OTPayrollError(
            "The attendance file has no valid rows after cleaning. Check the "
            "'Date' and 'Code' columns."
        )
    return df


# ── Optional employee master file (for Title / Department enrichment) ────────
def load_employee_master(file_obj):
    """
    Optional. Expected columns (case-insensitive, flexible):
      Code, (Employees Name | Name), Title, Department
    Returns {CodeKey: {"name": str, "title": str, "department": str}}
    Missing Title/Department columns are simply left as "—" for everyone.
    """
    if file_obj is None:
        return {}

    file_obj.seek(0)
    df = pd.read_excel(file_obj)
    df.columns = df.columns.str.strip()

    code_col = _find_col(df.columns, "Code")
    name_col = _find_col(df.columns, "Employees Name", "Name", "Employee Name")
    title_col = _find_col(df.columns, "Title", "Position", "Title / Position")
    dept_col = _find_col(df.columns, "Department", "Dept")

    if code_col is None:
        raise OTPayrollError("The employee data file needs a 'Code' column to match employees.")

    df["CodeKey"] = df[code_col].map(norm_code)
    df = df[df["CodeKey"].notna()]

    lookup = {}
    for _, row in df.iterrows():
        name = str(row[name_col]).strip() if name_col else ""
        title = str(row[title_col]).strip() if title_col else ""
        dept = str(row[dept_col]).strip() if dept_col else ""
        name = "" if name.lower() == "nan" else name
        title = "—" if (not title or title.lower() == "nan") else title
        dept = "—" if (not dept or dept.lower() == "nan") else dept
        lookup[row["CodeKey"]] = {"name": name, "title": title, "department": dept}
    return lookup


# ── Optional vacation file (for the Unpaid Leave check) ───────────────────────
def load_vacation_coverage(file_obj):
    """
    Optional. Any row of any type/status covering a date counts as
    'something exists' for that day, per the business rule: a day is only
    Unpaid Leave if there is NOTHING at all on record for it.

    Returns a dict: {CodeKey: [(from_date, to_date), ...]}
    """
    if file_obj is None:
        return {}

    file_obj.seek(0)
    df = pd.read_excel(file_obj)
    df.columns = df.columns.str.strip()

    code_col = _find_col(df.columns, "Code")
    from_col = _find_col(df.columns, "From", "From Date", "Start Date")
    to_col = _find_col(df.columns, "To", "To Date", "End Date")

    if code_col is None or from_col is None:
        raise OTPayrollError(
            "The vacation file needs at least 'Code' and 'From' columns to check leave coverage."
        )
    # If there's no explicit "To" column, treat each row as a single-day record.
    if to_col is None:
        df["_To"] = df[from_col]
        to_col = "_To"

    df["CodeKey"] = df[code_col].map(norm_code)
    df["_From"] = pd.to_datetime(df[from_col], errors="coerce").dt.date
    df["_To2"] = pd.to_datetime(df[to_col], errors="coerce").dt.date
    df = df.dropna(subset=["CodeKey", "_From", "_To2"])

    coverage = {}
    for _, row in df.iterrows():
        coverage.setdefault(row["CodeKey"], []).append((row["_From"], row["_To2"]))
    return coverage


def _is_covered(coverage, code_key, date):
    for from_date, to_date in coverage.get(code_key, []):
        if from_date <= date <= to_date:
            return True
    return False


# ── Holidays: manual list + optional uploaded file, merged ───────────────────
def parse_holiday_file(file_obj):
    """
    Optional holidays file. Looks for any column with 'date' in its name
    (case-insensitive); if there's only one column, uses that one instead.
    Returns a set of datetime.date.
    """
    if file_obj is None:
        return set()

    file_obj.seek(0)
    name = getattr(file_obj, "name", "")
    if name.lower().endswith(".csv"):
        df = pd.read_csv(file_obj)
    else:
        df = pd.read_excel(file_obj)
    df.columns = df.columns.str.strip()

    date_col = _find_col(df.columns, "Date", "Holiday Date", "Holiday")
    if date_col is None:
        if len(df.columns) == 1:
            date_col = df.columns[0]
        else:
            raise OTPayrollError(
                "Could not find a date column in the holidays file. "
                "Name the column 'Date' or upload a file with a single date column."
            )

    dates = pd.to_datetime(df[date_col], errors="coerce").dt.date.dropna().tolist()
    return set(dates)


# ── OT time-window math ───────────────────────────────────────────────────────
def _time_to_minutes(t):
    return t.hour * 60 + t.minute + t.second / 60.0


def _morning_overlap_minutes(a, b):
    """
    a, b are minutes on an extended timeline (b > a, can exceed 1440 for
    shifts that cross midnight). Returns total minutes overlapping any
    07:00–19:00 morning window, repeating every 1440 minutes.
    """
    total = 0.0
    k = int(a // 1440) - 1
    while True:
        win_start = 1440 * k + MORNING_START_MIN
        win_end = 1440 * k + MORNING_END_MIN
        if win_start >= b:
            break
        overlap_start = max(a, win_start)
        overlap_end = min(b, win_end)
        if overlap_end > overlap_start:
            total += overlap_end - overlap_start
        k += 1
    return total


def split_ot_hours(t_in, t_out, workday_hrs=DEFAULT_WORKDAY_HRS):
    """
    Given punch in/out (datetime.time) on a regular workday, returns
    (morning_ot_hours, night_ot_hours) for the overtime portion — i.e.
    everything worked beyond `workday_hrs` after t_in.

    Example: in=09:00, out=20:00 -> official ends 17:00, OT is 17:00-20:00 (3h)
             -> 17:00-19:00 = 2h morning OT, 19:00-20:00 = 1h night OT.
    """
    if not t_in or not t_out:
        return 0.0, 0.0

    in_min = _time_to_minutes(t_in)
    out_min = _time_to_minutes(t_out)
    if out_min < in_min:
        out_min += 24 * 60  # crossed midnight

    ot_start = in_min + workday_hrs * 60
    ot_end = out_min
    if ot_end <= ot_start:
        return 0.0, 0.0

    total_ot = ot_end - ot_start
    morning = _morning_overlap_minutes(ot_start, ot_end)
    night = total_ot - morning
    return round(morning / 60, 2), round(night / 60, 2)


def calc_hours(t_in, t_out):
    if not t_in or not t_out:
        return None
    a = datetime.datetime(2000, 1, 1, t_in.hour, t_in.minute, t_in.second)
    b = datetime.datetime(2000, 1, 1, t_out.hour, t_out.minute, t_out.second)
    if b < a:
        b += datetime.timedelta(days=1)
    return round((b - a).seconds / 3600, 2)


# ── Main report builder ───────────────────────────────────────────────────────
def generate_ot_payroll_report(
    attendance_file,
    employee_file=None,
    vacation_file=None,
    holiday_dates=None,
    weekend_days=None,
    workday_hrs=DEFAULT_WORKDAY_HRS,
    progress_callback=None,
):
    """
    Returns (report_df, stats) where report_df has exactly these columns:
      Name, Title / Position, Department, Morning OT Hours, Night OT Hours,
      Cancel Day Offs / Days, Official Holiday / Days, Unpaid Leave,
      Total Working Days
    """
    if weekend_days is None:
        weekend_days = DEFAULT_WEEKEND_DAYS
    if holiday_dates is None:
        holiday_dates = set()

    df_att = load_attendance(attendance_file)
    emp_lookup = load_employee_master(employee_file)
    vac_coverage = load_vacation_coverage(vacation_file)

    date_min = df_att["Date"].min()
    date_max = df_att["Date"].max()
    all_dates = [date_min + datetime.timedelta(days=i) for i in range((date_max - date_min).days + 1)]

    codes = sorted(df_att["CodeKey"].unique())
    rows = []
    total_codes = len(codes)

    for idx, code_key in enumerate(codes, 1):
        person_df = df_att[df_att["CodeKey"] == code_key]

        # Name resolution: prefer employee master, fall back to attendance file
        master = emp_lookup.get(code_key, {})
        name = master.get("name") or (
            str(person_df["Name"].iloc[0]).strip() if "Name" in person_df.columns else ""
        )
        title = master.get("title", "—")
        department = master.get("department", "—")

        # Daily in/out per date
        daily = {}
        for date, g in person_df.groupby("Date"):
            ins = g[g["I/O"] == "Punch In"]["Time"].dropna().tolist()
            outs = g[g["I/O"] == "Punch Out"]["Time"].dropna().tolist()
            daily[date] = (min(ins) if ins else None, max(outs) if outs else None)

        morning_ot_total = 0.0
        night_ot_total = 0.0
        cancel_day_off_days = 0.0
        official_holiday_days = 0.0
        unpaid_leave_days = 0
        working_days = 0

        for date in all_dates:
            t_in, t_out = daily.get(date, (None, None))
            has_punch_in = t_in is not None
            hrs = calc_hours(t_in, t_out)

            if has_punch_in:
                working_days += 1

            is_special = (date.weekday() in weekend_days) or (date in holiday_dates)

            if is_special:
                if hrs:
                    if hrs >= workday_hrs:
                        official_holiday_days += 1.0
                    else:
                        cancel_day_off_days += hrs / workday_hrs
                # no OT and no unpaid-leave check on weekend/holiday days
                continue

            # Regular day: OT calculation
            m_ot, n_ot = split_ot_hours(t_in, t_out, workday_hrs)
            morning_ot_total += m_ot
            night_ot_total += n_ot

            # Unpaid leave: nothing at all on record for this date
            if not has_punch_in and not _is_covered(vac_coverage, code_key, date):
                unpaid_leave_days += 1

        rows.append(
            {
                "Name": name,
                "Title / Position": title,
                "Department": department,
                "Morning OT Hours": round(morning_ot_total, 2),
                "Night OT Hours": round(night_ot_total, 2),
                "Cancel Day Offs / Days": round(cancel_day_off_days, 2),
                "Official Holiday / Days": round(official_holiday_days, 2),
                "Unpaid Leave": unpaid_leave_days,
                "Total Working Days": working_days,
            }
        )

        if progress_callback:
            progress_callback(idx, total_codes)

    report_df = pd.DataFrame(
        rows,
        columns=[
            "Name",
            "Title / Position",
            "Department",
            "Morning OT Hours",
            "Night OT Hours",
            "Cancel Day Offs / Days",
            "Official Holiday / Days",
            "Unpaid Leave",
            "Total Working Days",
        ],
    ).sort_values("Name").reset_index(drop=True)

    stats = {
        "num_employees": len(report_df),
        "num_days": len(all_dates),
        "date_min": date_min,
        "date_max": date_max,
    }
    return report_df, stats


# ── Excel export, styled to match the app's look ──────────────────────────────
_THIN = Side(style="thin")
_BDR = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F2F6FC")


def export_report_to_excel(report_df, title_text="OT & Payroll Summary"):
    """Returns an openpyxl Workbook with the report nicely formatted."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OT & Payroll Summary"

    columns = list(report_df.columns)
    col_widths = [22, 20, 18, 16, 14, 18, 18, 12, 16]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Arial", bold=True, size=13, color="1F3864")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    header_row = 2
    for col_idx, col_name in enumerate(columns, 1):
        c = ws.cell(row=header_row, column=col_idx, value=col_name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[header_row].height = 30

    for i, (_, row) in enumerate(report_df.iterrows()):
        r = header_row + 1 + i
        fill = _ALT_FILL if i % 2 == 0 else None
        for col_idx, col_name in enumerate(columns, 1):
            val = row[col_name]
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = _BDR
            c.alignment = Alignment(
                horizontal="left" if col_name in ("Name", "Title / Position", "Department") else "center",
                vertical="center",
            )
            if fill:
                c.fill = fill

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(report_df)}"
    return wb
